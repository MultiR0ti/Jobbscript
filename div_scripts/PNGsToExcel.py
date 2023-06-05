

import pandas as pd
#import xlsxwriter
import openpyxl
from openpyxl import workbook
import os




def make_columns(ws, depth_dict, bp_name):
    column_header_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor='ffbdbcb9')
    ws['A1'] = 'Borhull'
    ws['B1'] = 'Start'
    ws['C1'] = 'End'
    ws['D1'] = 'Geology'
    ws['E1'] = 'Kvalitet'
    ws['F1'] = 'Max_depth:'
    ws['G1'] = depth_dict.get(bp_name)

    for col in range(1, 6):
        ws.cell(1, col).fill = column_header_fill

    #cell_format = workbook.add_format()
    #cell_format.set_bg_color('#c0c0c0')
    #ws.set_column('A:E', 1, cell_format)


def make_rows(ws, bornr: str):
    #ws.write('A2', bornr)
    #ws.write('A3', bornr)
    #ws.write('B2', 0.0)
    ws['A2'] = bornr
    ws['A3'] = bornr
    ws['B2'] = 0.0
    ws['B3'] = '=C2'
    ws['B4'] = '=C3'
    ws['B5'] = '=C4'
    ws['B6'] = '=C5'

def create_depth_dict(wb):
    
    ws = wb['GS_Database']
    bp_dict = {}
    row = 2
    while ws.cell(row, 1):
        print(ws.cell(row, 1).internal_value)
        if not ws.cell(row, 1).internal_value:
            break
        print(ws.cell(row, 7).internal_value)
        print(ws.cell(row, 8).internal_value, type(ws.cell(row, 8).internal_value))
        # If the cell is not empty and is a string
        if ws.cell(row, 8).internal_value and isinstance(ws.cell(row, 8).internal_value, str):
            bp_dict[ws.cell(row, 1).internal_value] = float(ws.cell(row, 7).internal_value.replace(',', '.')) + float(ws.cell(row, 8).internal_value.replace(',', '.'))
        # If the cell is not empty and its not a string
        elif ws.cell(row, 8).internal_value and not isinstance(ws.cell(row, 8).internal_value, str):
            bp_dict[ws.cell(row, 1).internal_value] = ws.cell(row, 7).internal_value + ws.cell(row, 8).internal_value
        # If the cell is empty and its a string
        elif ws.cell(row, 8).internal_value is None and isinstance(ws.cell(row, 7).internal_value, str):
            bp_dict[ws.cell(row, 1).internal_value] = float(ws.cell(row, 7).internal_value.replace(',', '.'))
        else:
            bp_dict[ws.cell(row, 1).internal_value] = ws.cell(row, 7).internal_value
        row +=1
    return bp_dict


def main():
    '''Create a sheet in excel with png of boring 
    profile and create interpretation structure'''

    img_path = r'\\nsv2-nasuni-02\fredrikstad\Prosjekt\O10245\10245026-01\10245026-01-03_ARBEIDSOMRAADE\21_fagomraade\11_Geoteknikk\10245026-10-GEOSUITE\Supplerende GRUS 2023\AUTOGRAF.RIT\images'
    excel_file = r'C:\Users\jdr\ProgrammingDocuments\Tolkning_Sarpsbru_SB-Boringer_onedrive_manglerstreker.xlsm'

    # Create an new Excel file and add a worksheet.
    wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)

    dept_dict = create_depth_dict(wb)

    cpt_pngs = []

    for png in sorted(os.listdir(img_path)):
        print(png)
        #fullname = png.basename()
        #filename = fullname.splitext()
        bp_name = png.split('.')[0]

        if 'C' in bp_name:
            cpt_pngs.append(img_path + r'\\' + png)
            continue

        if png.startswith('SB') and 'PR' not in png and 'RIG-TEG' not in png and bp_name not in wb.sheetnames and 'C' not in bp_name:
            ws = wb.create_sheet(bp_name)
            bp_png = openpyxl.drawing.image.Image(img_path + r'\\' + png)
            bp_png.anchor = 'F3'
            ws.add_image(bp_png)
            make_columns(ws, dept_dict, bp_name)
            make_rows(ws, bp_name)

    '''for cpt in cpt_pngs:
        try:
            ws = wb['SB-' + cpt.split('B')[-1].split('C')[0]]
        except Exception as e:
            ws = wb['SB-' + cpt.split('B-')[-1].split('C')[0]]

        cpt_png = openpyxl.drawing.image.Image(cpt_pngs[0])
        cpt_png.anchor = 'AA3'
        ws.add_image(cpt_png)'''

    # wb._sheets.sort(key=lambda ws: ws.title)
    wb.save(excel_file)

if __name__ == "__main__":
    main()